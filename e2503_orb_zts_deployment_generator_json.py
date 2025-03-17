import os
import re
import json
from openpyxl import load_workbook

def log(msg, log_lines):
    print(msg)
    log_lines.append(msg)

def load_json_template(json_template_path, log_lines):
    if not os.path.exists(json_template_path):
        log(f"JSON template file not found at: {json_template_path}", log_lines)
        return None
    with open(json_template_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data

def save_json(data, output_path, log_lines):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)
    log(f"Generated JSON configuration saved to: {output_path}", log_lines)

def save_log(log_lines, output_log_path):
    with open(output_log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))
    print(f"Log file saved to: {output_log_path}")

def save_var(dp_data, output_var_path, log_lines):
    with open(output_var_path, "w", encoding="utf-8") as f:
        json.dump(dp_data, f, indent=4)
    log(f"Variable file saved to: {output_var_path}", log_lines)

def sanitize_key(key):
    return key.strip()

def normalize_key(key):
    # Lower-case and remove non-alphanumeric characters for robust matching.
    return re.sub(r'[^a-z0-9]', '', key.lower())

def convert_value(key, val):
    if val is None:
        return ""
    # For AIF_Enabled always preserve string from Excel.
    if key == "AIF_Enabled":
        return str(val).strip()
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return ""
        lower_s = s.lower()
        if lower_s == "true":
            return True
        elif lower_s == "false":
            return False
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s) if '.' in s else int(s)
        return s
    return val

def read_excel_dp_data(excel_path, site, log_lines):
    if not os.path.exists(excel_path):
        log(f"Excel file not found at: {excel_path}", log_lines)
        return {}
    
    wb = load_workbook(filename=excel_path, data_only=True)
    if "DP" not in wb.sheetnames:
        log("Sheet 'DP' not found in the Excel file.", log_lines)
        return {}
    
    ws = wb["DP"]
    header = [cell.value.strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        key_index = header.index("NE Parameter Name")
    except ValueError:
        log("Column 'NE Parameter Name' not found in header.", log_lines)
        return {}
    
    site_index = None
    for idx, col_header in enumerate(header):
        if col_header.lower() == site.lower():
            site_index = idx
            break
    if site_index is None:
        log(f"Site column '{site}' not found in header.", log_lines)
        return {}
    
    dp_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[key_index]
        if key is None:
            continue
        key = sanitize_key(str(key))
        raw_val = row[site_index]
        value = "" if raw_val is None else (str(raw_val).strip() if key == "AIF_Enabled" else convert_value(key, raw_val))
        dp_data[key] = value
        log(f"DP data: Set key '{key}' to value: {value} (type: {type(value).__name__})", log_lines)
    return dp_data

def build_dp_lookup(dp_data):
    dp_lookup = {}
    for key, value in dp_data.items():
        dp_lookup[normalize_key(key)] = value
    return dp_lookup

def recursive_update_template(item, dp_lookup, log_lines):
    if isinstance(item, dict):
        if "name" in item and "value" in item:
            param_name = item["name"]
            norm_param = normalize_key(param_name)
            if norm_param in dp_lookup:
                old_val = item["value"]
                new_val = dp_lookup[norm_param]
                if not isinstance(old_val, (dict, list)):
                    item["value"] = new_val
                    log(f"Updated parameter '{param_name}' from {old_val} to {new_val}", log_lines)
            recursive_update_template(item["value"], dp_lookup, log_lines)
        else:
            for key, value in item.items():
                norm_key = normalize_key(key)
                if norm_key in dp_lookup and not isinstance(value, (dict, list)):
                    old_val = item[key]
                    new_val = dp_lookup[norm_key]
                    item[key] = new_val
                    log(f"Updated key '{key}' from {old_val} to {new_val}", log_lines)
                else:
                    recursive_update_template(value, dp_lookup, log_lines)
    elif isinstance(item, list):
        for sub_item in item:
            recursive_update_template(sub_item, dp_lookup, log_lines)

def merge_into_composite(composite_obj, prefix, dp_lookup, log_lines):
    norm_prefix = normalize_key(prefix)
    existing_keys = {normalize_key(k) for k in composite_obj.keys()}
    for key_norm, value in dp_lookup.items():
        if key_norm.startswith(norm_prefix) and key_norm not in existing_keys:
            composite_obj[key_norm] = value
            log(f"Merged key '{key_norm}' into composite section under prefix '{prefix}': {value}", log_lines)

def update_json_template(json_template, dp_data, log_lines):
    dp_lookup = build_dp_lookup(dp_data)
    recursive_update_template(json_template, dp_lookup, log_lines)
    if isinstance(json_template, list):
        for obj in json_template:
            if isinstance(obj, dict) and obj.get("name") == "chartValues" and isinstance(obj.get("value"), dict):
                chart_values = obj["value"]
                if "UmIdpConfig" in chart_values and isinstance(chart_values["UmIdpConfig"], dict):
                    merge_into_composite(chart_values["UmIdpConfig"], "UmIdp", dp_lookup, log_lines)
    return json_template

def main():
    # Gather inputs.
    excel_path = input("Enter the path to the Excel file (DP sheet is used): ").strip()
    json_template_path = input("Enter the path to the JSON Template file (e.g., zts_template.json): ").strip()
    sites_str = input("Enter the SITE column names separated by comma: ").strip()
    sites = [s.strip() for s in sites_str.split(",") if s.strip() != ""]
    output_dir = input("Enter the output directory path (where output files will be saved): ").strip()
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    json_template = load_json_template(json_template_path, [])
    if json_template is None:
        return

    for site in sites:
        log_lines = []
        log(f"Processing site: {site}", log_lines)
        dp_data = read_excel_dp_data(excel_path, site, log_lines)
        if not dp_data:
            log(f"No valid DP data found for site: {site}. Skipping.", log_lines)
            continue
        updated_json = update_json_template(json_template, dp_data, log_lines)
        
        # Create output file names following <SITE>_zts_24.7_mp1.*
        base_filename = f"{site}_zts_24.7_mp1"
        output_json_filename = os.path.join(output_dir, base_filename + ".json")
        output_log_filename = os.path.join(output_dir, base_filename + ".log")
        output_var_filename = os.path.join(output_dir, base_filename + ".var")
        
        save_json(updated_json, output_json_filename, log_lines)
        save_log(log_lines, output_log_filename)
        save_var(dp_data, output_var_filename, log_lines)

if __name__ == "__main__":
    main()